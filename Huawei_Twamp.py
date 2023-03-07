import pandas as pd
from tkinter import filedialog
import openpyxl as op
import xlsxwriter
import os
from openpyxl.utils.dataframe import dataframe_to_rows
from tkinter import *

print('Developed by Shekh Mohammad Nasim')
print("If any support required, feel free to contact")
print("'shekh.nasim@gmail.com'")
# brows the file and store the complete directory in a variable
def openFile():
    input_file_name = filedialog.askopenfilename(filetypes=(('Excel files','xls'), ('All Files', '.*')))
    return input_file_name

input_file_name = openFile()
input_file_directory = '/'.join(input_file_name.split('/')[0:-1]) # Store the folder directory in a variable

# Create dataframe for input and output files
df_input = pd.read_excel(input_file_name)

# Create output Excel file if not exist
filelist = os.listdir(input_file_directory) #Store the file as a list in a filelist variable
if 'Huawei Twamp.xlsx' in filelist:
    print("Huawei Twamp.xlsx file exist")
    print('Processing.......Please wait......')
else:
    print("Huawei Twamp.xlsx file does not exist, now creating the file")
    wb = xlsxwriter.Workbook(input_file_directory+'/'+'Huawei Twamp.xlsx') # Create an workbook in the given directory and the given name
    ws1 = wb.add_worksheet('Huawei_Twamp_Delay') # Create an worksheet in the workbook
    ws2 = wb.add_worksheet('Huawei_Twamp_Jitter') # Create an worksheet in the workbook
    ws3 = wb.add_worksheet('Huawei_Twamp_Packet_Loss') # Create an worksheet in the workbook
    ws4 = wb.add_worksheet('Huawei_Twamp_Violation_count') # Create an worksheet in the workbook

    #ws_output_temp = wb_output.add_worksheet('TEMP') # Create an worksheet in the workbook
    ws1.write(0,0,'site_id') # write text in A1 cell/0,0 location
    ws2.write(0,0,'site_id') # write text in A1 cell/0,0 location
    ws3.write(0,0,'site_id') # write text in A1 cell/0,0 location
    ws4.write(0,0,'site_id') # write text in A1 cell/0,0 location
    wb.close()
    print('Processing.......Please wait......')

# Load the output workbook
wb = op.load_workbook(input_file_directory +'/'+'Huawei Twamp.xlsx')
ws1 = wb['Huawei_Twamp_Delay']
ws2 = wb['Huawei_Twamp_Jitter']
ws3 = wb['Huawei_Twamp_Packet_Loss']
ws4 = wb['Huawei_Twamp_Violation_count']

df1=pd.read_excel(input_file_directory +'/'+'Huawei Twamp.xlsx', sheet_name="Huawei_Twamp_Delay")
df2=pd.read_excel(input_file_directory +'/'+'Huawei Twamp.xlsx', sheet_name="Huawei_Twamp_Jitter")
df3=pd.read_excel(input_file_directory +'/'+'Huawei Twamp.xlsx', sheet_name="Huawei_Twamp_Packet_Loss")
df4=pd.read_excel(input_file_directory +'/'+'Huawei Twamp.xlsx', sheet_name="Huawei_Twamp_Violation_count")

#VLookup
Huawei_Twamp_Delay=pd.merge(df1,df_input, how='outer', on='site_id')
Huawei_Twamp_Jitter=pd.merge(df2,df_input, how='outer', on='site_id')
Huawei_Twamp_Packet_Loss=pd.merge(df3,df_input, how='outer', on='site_id')
Huawei_Twamp_Violation_count=pd.merge(df4,df_input, how='outer', on='site_id')

#first converting string to date, 2nd formating the date as required
#import datetime as dt
Huawei_Twamp_Delay['Date']=pd.to_datetime(Huawei_Twamp_Delay['Date'], format='%Y/%m/%d')
Huawei_Twamp_Delay['Date']=Huawei_Twamp_Delay['Date'].dt.strftime('%Y/%m/%d')
dd=Huawei_Twamp_Delay['Date'].iloc[0] # Get first row of a given column.

dd=Huawei_Twamp_Delay['Date'].iloc[0] # Get first row of a given column.

# drop unwanted columns
Twamp_Delay = Huawei_Twamp_Delay.drop(['Date','jitter','packet_loss','violation_count'], axis=1)
Twamp_Jitter = Huawei_Twamp_Jitter.drop(['Date','delay','packet_loss','violation_count'], axis=1)
Twamp_Packet_Loss = Huawei_Twamp_Packet_Loss.drop(['Date','jitter','delay','violation_count'], axis=1)
Twamp_Violation_count = Huawei_Twamp_Violation_count.drop(['Date','jitter','packet_loss','delay'], axis=1)

#Rename Header
Twamp_Delay.rename(columns={'delay':dd}, inplace=True)
Twamp_Jitter.rename(columns={'jitter':dd}, inplace=True)
Twamp_Packet_Loss.rename(columns={'packet_loss':dd}, inplace=True)
Twamp_Violation_count.rename(columns={'violation_count':dd}, inplace=True)

# Delete all the rows in the worksheet
ws1.delete_rows(1, ws1.max_row)
ws2.delete_rows(1, ws2.max_row)
ws3.delete_rows(1, ws3.max_row)
ws4.delete_rows(1, ws4.max_row)

# Write the modified data back to the worksheet
for row in dataframe_to_rows(Twamp_Delay, index=False, header=True):
    ws1.append(row)
for row in dataframe_to_rows(Twamp_Jitter, index=False, header=True):
    ws2.append(row)
for row in dataframe_to_rows(Twamp_Packet_Loss, index=False, header=True):
    ws3.append(row)
for row in dataframe_to_rows(Twamp_Violation_count, index=False, header=True):
    ws4.append(row)
wb.save(input_file_directory+'/'+'Huawei Twamp.xlsx')
wb.close()
print('Processing completed........')